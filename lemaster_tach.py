# https://www.youtube.com/watch?v=718edSNvKLA
import pandas as pd
import openpyxl as xl

"""
Excel 파일 읽기 및 쓰기
"""

# Load the excel file
df = pd.read_excel('./data/menu.xlsx', engine='openpyxl')

workbook = xl.load_workbook('./data/menu.xlsx')
worksheet = workbook.active

# worksheet에서 문자로 구성된 열은 건너뛰고, 숫자로 된 열만 찾아, 해당 열 맨 아래열 열의 합계를 추가
for col in worksheet.iter_cols(min_row=2, max_row=8, min_col=2, max_col=worksheet.max_column):
    numeric_col = True
    col_sum = 0
    for cell in col:
        if isinstance(cell.value, (int, float)):
            col_sum += cell.value
        else:
            numeric_col = False
            break

    if numeric_col:
        worksheet.cell(row=9, column=cell.column, value=col_sum)


"""
Style 
"""
# 전체 셀의 값 출력
for row in range(1,4):
    for col in range(1,4):
        print(worksheet.cell(row=row, column=col).value)

for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)

# Apply style to header
for row in worksheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.fill = xl.styles.PatternFill(start_color='FFC000', end_color='00C000', fill_type='solid')
        cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
        cell.font = xl.styles.Font(size=16, bold=True)


# Auto adjust the cell width
# 행과 열 반복
for row in worksheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
    print(row)

for row in worksheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=1):
    print(row)

# 헤더에 색상 추가
for rowcell in worksheet.iter_rows(min_row=1, max_row=1): 
    for cell in rowcell:
        cell.fill      = xl.styles.PatternFill(start_color='FFC000', end_color='00C000', fill_type='solid')
        cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
        cell.font      = xl.styles.Font(size=16, bold=True)

# 셀너비를 자동으로 조정
worksheet.column_dimensions['A'].width = 20
worksheet.row_dimensions[1].height = 30

for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column].width = adjusted_width  

# 차트, 이미지, 피벗 테이블 등 추가: openpyxl은 차트, 이미지, 피벗 테이블 등을 워크시트에 추가하는 기능을 제공합니다. 이를 통해 데이터를 시각화하고 보고서를 더욱 풍부하게 만들 수 있습니다.
# 차트 추가
chart = xl.chart.BarChart()



# 데이터에 윤곽선 추가
for rowcell in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    for cell in rowcell:
        cell.border = xl.styles.Border(
            left  =xl.styles.Side(style='thin'), 
            right =xl.styles.Side(style='thin'), 
            top   =xl.styles.Side(style='thin'), 
            bottom=xl.styles.Side(style='thin')
        )



workbook.save('./data/menu_result.xlsx')